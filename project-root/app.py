import csv
import uuid
from flask import Flask, render_template, request, redirect, url_for, jsonify
import os
import pandas as pd
import sqlite3
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


app = Flask(__name__)

# Configuración
app.config['UPLOAD_FOLDER'] = 'uploads/'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
DATABASE_PATH = 'mi_base_local.db'
ALLOWED_EXTENSIONS = {'csv', 'xls', 'xlsx'}

conn = sqlite3.connect('mi_base_local.db')
conn.close()

@app.route('/')
def index():
    return render_template('index.html')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xls', 'xlsx', 'csv'}

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'files[]' not in request.files:
            return "No file part"
        files = request.files.getlist('files[]')
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                process_and_upload_to_sqlite(filepath)
        return redirect(url_for('list_tables'))
    return render_template('upload.html')

def process_and_upload_to_sqlite(filepath):
    try:
        if filepath.endswith('.csv'):
            data = pd.read_csv(filepath)
        elif filepath.endswith(('.xls', '.xlsx')):
            workbook = load_workbook(filepath, data_only=True)
            sheet = workbook.active
            rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
            data = pd.DataFrame(rows)
        else:
            return

        table_name = os.path.splitext(os.path.basename(filepath))[0]
        with sqlite3.connect(DATABASE_PATH) as conn:
            data.to_sql(table_name, conn, if_exists='replace', index=False)
    except Exception as e:
        print(f"Error al procesar el archivo {filepath}: {e}")

"""
def process_and_upload_to_sqlite(filepath):
    try:
        # leemos el archivo que puede ser tipo csv o excel
        if filepath.endswith('.csv'):
            data = pd.read_csv(filepath)
        elif filepath.endswith(('.xls', '.xlsx')):
            data = pd.read_excel(filepath, engine='openpyxl')
        else:
            print(f"Formato no soportado: {filepath}")
            return

        # conecto a SQLite
        conn = sqlite3.connect(DATABASE_PATH)
        table_name = os.path.splitext(os.path.basename(filepath))[0]

        print(f"Creando tabla: {table_name}")
        # guardo los datos en SQLite
        data.to_sql(table_name, conn, if_exists='replace', index=False)
        
        # verifico las tablas existentes
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        print(f"Tablas existentes: {tables}")

        conn.close()
        print(f"Datos cargados en SQLite desde {filepath}")

        print(f"Vista previa de los datos:\n{data.head()}")
    except Exception as e:
        print(f"Error al procesar el archivo {filepath}: {e}")
"""
"""
def process_and_upload_to_sqlite(filepath):
    try:
        if filepath.endswith(('.xls', '.xlsx')):
            workbook = load_workbook(filepath, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                fill_merged_cells(sheet)

                # Leer datos de la hoja como DataFrame
                data = pd.DataFrame(sheet.values)

                # Configuración adicional: rellenar columnas vacías
                data = data.fillna(method='ffill', axis=0)  # Hacia abajo
                data = data.fillna(method='ffill', axis=1)  # Hacia la derecha

                # Guardar en SQLite
                table_name = f"{os.path.splitext(os.path.basename(filepath))[0]}_{sheet_name}"
                conn = sqlite3.connect(DATABASE_PATH)
                data.to_sql(table_name, conn, if_exists='replace', index=False)
                print(f"Datos de {sheet_name} guardados correctamente en la tabla {table_name}")
                conn.close()
    except Exception as e:
        print(f"Error al procesar el archivo {filepath}: {e}")
"""
@app.route('/validate_sample', methods=['POST'])
def validate_sample():
    try:
        #obtenemos los datos del formulario
        population = int(request.form['population'])
        #sample_size = int(request.form['sample_size'])

        #CALCULO VALIDEZ MUESTRA ??
        if (population*(1.64485**2)*0.5*0.5)/(((population-1)*(0.1**2))+((1.64485**2)*0.5*0.5))>10:
            result = "La muestra es válida para representar a la población."
        else:
            result = "La muestra no es válida, debe ser XXXXXXXXXXXX"

        return render_template('upload.html', validation_result=result)
    except ValueError:
        return render_template('upload.html', validation_result="Por favor, introduzca valores numéricos válidos.")


@app.route('/tables')
def list_tables():
    try:
        with sqlite3.connect(DATABASE_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [row[0] for row in cursor.fetchall()]
            print(f"Tablas en la base de datos: {tables}")
        return render_template('tables.html', tables=tables)
    except Exception as e:
        return f"Error al listar tablas: {e}"
list_tables()

@app.route('/table/<table_name>')
def show_table(table_name):
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
        conn.close()
        return render_template('table.html', table_name=table_name, data=df.to_html(index=False))
    except Exception as e:
        return f"Error al mostrar la tabla {table_name}: {e}"

@app.route('/delete_table/<table_name>', methods=['POST'])
def delete_table(table_name):
    try:
        # Validar el nombre de la tabla
        if not table_name.isidentifier():
            return f"Nombre de tabla no válido: {table_name}", 400

        # Conectar a la base de datos
        with sqlite3.connect(DATABASE_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
            conn.commit()

        return redirect(url_for('list_tables'))  # Redirige a la lista de tablas
    except Exception as e:
        return f"Error al eliminar la tabla {table_name}: {e}", 500

@app.route('/get_columns', methods=['GET'])
def get_columns():
    table_name = request.args.get('table')
    try:
        with sqlite3.connect(DATABASE_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute(f"PRAGMA table_info({table_name});")
            columns = [col[1] for col in cursor.fetchall()]
        return jsonify(columns)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/join_tables', methods=['GET', 'POST'])
def join_tables():
    if request.method == 'POST':
        # Lógica para procesar la unión de tablas
        table1 = request.form.get('table1')
        column1 = request.form.get('column1')
        table2 = request.form.get('table2')
        column2 = request.form.get('column2')

        if not (table1 and column1 and table2 and column2):
            return "Por favor, seleccione todas las opciones.", 400

        try:
            with sqlite3.connect(DATABASE_PATH) as conn:
                query = f"""
                    SELECT * FROM {table1}
                    INNER JOIN {table2}
                    ON {table1}.{column1} = {table2}.{column2}
                """
                df = pd.read_sql_query(query, conn)
            return render_template('joined_tables.html', data=df.to_html(index=False))
        except Exception as e:
            return f"Error al relacionar tablas: {e}", 500

    # Lógica para manejar la carga inicial del formulario (GET)
    try:
        with sqlite3.connect(DATABASE_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [row[0] for row in cursor.fetchall()]

            # Obtener columnas para cada tabla
            columns = {}
            for table in tables:
                cursor.execute(f"PRAGMA table_info({table});")
                columns[table] = [col[1] for col in cursor.fetchall()]

        return render_template('join_tables.html', tables=tables, columns=columns)
    except Exception as e:
        return f"Error al cargar las tablas: {e}"


def fill_merged_cells(sheet):
    """Rellena las celdas combinadas con el valor principal."""
    for merged_cell_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell_range))
        top_left_value = sheet.cell(row=min_row, column=min_col).value

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if sheet.cell(row=row, column=col).value is None:
                    sheet.cell(row=row, column=col).value = top_left_value


@app.route('/paste_tables', methods=['GET'])
def paste_tables():
    return render_template('paste_tables.html')

"""
@app.route('/save_pasted_data', methods=['POST'])
def save_pasted_data():
    try:
        data = request.json.get('data', '')
        table_name = request.json.get('tableName', '').strip()

        if not data or not all(isinstance(row, (list, tuple)) for row in data):
            print(f"Error: Los datos en {table_name} no son válidos.")
            return

        if not table_name:
            return "Error: No se proporcionó un nombre para la tabla", 400

        # Validar el nombre de la tabla
        if not table_name.isidentifier():
            return "Error: El nombre de la tabla contiene caracteres no válidos", 400

        # Procesar filas y columnas
        rows = [row.split('\t') for row in data.split('\n') if row.strip()]
        print(f"Filas originales: {rows}")
        #rows = [row.split('\t') for row in data.split('\n') if row.strip()]
        #rows = [[cell.strip() if isinstance(cell, str) else cell for cell in row] for row in rows]
        
        rows = fill_combined_cells(rows)
        print(f"Filas procesadas tras rellenar celdas combinadas: {rows}")

        # Validar consistencia de columnas
        max_columns = max(len(row) for row in rows)
        rows = [row + [""] * (max_columns - len(row)) for row in rows]

        # Sanitizar encabezados
        headers = rows[0]
        if not all(header.strip() for header in headers):
            headers = [f"Columna_{i+1}" for i in range(max_columns)]
            rows.insert(0, headers)
        else:
            headers = [col.strip().replace(" ", "_").replace("-", "_") for col in headers]

        print(f"Encabezados: {headers}")

        # Crear DataFrame
        df = pd.DataFrame(rows[1:], columns=headers)
        print(f"DataFrame creado:\n{df.head()}")

        # Guardar en CSV
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{table_name}.csv")
        df.to_csv(filepath, index=False, encoding='utf-8')
        print(f"Archivo CSV guardado correctamente en: {filepath}")

        # Guardar en SQLite
        with sqlite3.connect(DATABASE_PATH) as conn:
            df.to_sql(table_name, conn, if_exists='replace', index=False)

        return f"Datos guardados correctamente en la tabla '{table_name}' con {len(df)} filas.", 200
    except Exception as e:
        return f"Error al guardar los datos: {e}", 500

def extract_tables_with_merged_cells(filepath):
    workbook = load_workbook(filepath)
    tables = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Extraer datos de la hoja
        table_data = []
        for row in sheet.iter_rows(values_only=True):
            # Manejar celdas basadas en su tipo
            row_data = [(str(cell).strip() if isinstance(cell, str) else cell) for cell in row]
            table_data.append(row_data)

        # Debugging para verificar las primeras filas
        print(f"Datos extraídos de la hoja '{sheet_name}': {table_data[:5]}")

        # Procesar celdas combinadas (si aplica)
        table_data = fill_combined_cells(table_data)

        tables.append({
            "sheet_name": sheet_name,
            "data": table_data
        })

    return tables
"""

"""
@app.route('/save_pasted_data', methods=['POST'])
def save_pasted_data():
    try:
        # Obtener datos del cliente
        data = request.json.get('data', '').strip()
        table_name = request.json.get('tableName', '').strip()

        if not data or not table_name:
            return "Error: Datos o nombre de tabla no proporcionados", 400

        # Validar el nombre de la tabla
        if not table_name.isidentifier():
            return "Error: El nombre de la tabla contiene caracteres no válidos", 400

        # Procesar filas y columnas
        rows = [row.split('\t') for row in data.split('\n') if row.strip()]
        if not rows:
            return "Error: Los datos están vacíos o mal formateados", 400

        # Validar consistencia de columnas
        max_columns = max(len(row) for row in rows)
        rows = [row + [""] * (max_columns - len(row)) for row in rows]

        # Sanitizar encabezados
        headers = rows[0]
        if not all(isinstance(header, str) and header.strip() for header in headers):
            headers = [f"Columna_{i+1}" for i in range(max_columns)]

        # Crear DataFrame
        df = pd.DataFrame(rows[1:], columns=headers)
        print(f"DataFrame creado:\n{df.head()}")

        # Guardar en SQLite
        with sqlite3.connect(DATABASE_PATH) as conn:
            df.to_sql(table_name, conn, if_exists='replace', index=False)

        return f"Datos guardados correctamente en la tabla '{table_name}' con {len(df)} filas.", 200
    except Exception as e:
        print(f"Error al guardar los datos: {e}")
        return f"Error al guardar los datos: {e}", 500
"""
@app.route('/save_pasted_data', methods=['POST'])
def save_pasted_data():
    try:
        data = request.json.get('data', '').strip()
        table_name = request.json.get('tableName', '').strip()

        if not data or not table_name:
            return "Error: Datos o nombre de tabla no proporcionados", 400

        rows = [row.split('\t') for row in data.split('\n') if row.strip()]
        max_columns = max(len(row) for row in rows)
        rows = [row + [""] * (max_columns - len(row)) for row in rows]

        headers = rows[0]
        if not all(headers):
            headers = [f"Columna_{i+1}" for i in range(max_columns)]

        df = pd.DataFrame(rows[1:], columns=headers)
        with sqlite3.connect(DATABASE_PATH) as conn:
            df.to_sql(table_name, conn, if_exists='replace', index=False)

        return f"Datos guardados en '{table_name}'.", 200
    except Exception as e:
        return f"Error al guardar los datos: {e}", 500

"""
@app.route('/upload_excel', methods=['GET', 'POST'])
def upload_excel():
    if request.method == 'POST':
        # Guardar el archivo temporalmente
        file = request.files['file']
        file_path = f"/tmp/{file.filename}"
        file.save(file_path)

        # Procesar el archivo subido
        #tables = extract_tables_with_merged_cells(file_path)

        # Renderizar una vista previa
        return render_template('tables_preview.html', tables=tables)

    return render_template('upload_excel.html')
"""

if __name__ == '__main__':
    app.run(debug=True)